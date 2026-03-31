import openpyxl
import csv

xlsx_file = "inventario.xlsx"
csv_file = "inventario.csv"

wb = openpyxl.load_workbook(xlsx_file)
ws = wb.active

# Buscar la fila donde realmente empiezan los encabezados
header_row = None
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row and "ID de inventario" in [str(c).strip() if c else "" for c in row]:
        header_row = i
        break

if not header_row:
    raise Exception("No se encontró la fila de encabezados.")

# Extraer solo la tabla real
rows = list(ws.iter_rows(values_only=True))[header_row - 1:]

with open(csv_file, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)

    for row in rows:
        clean_row = []
        for cell in row:
            if cell is None:
                clean_row.append("")
            else:
                clean_row.append(str(cell).replace("\n", " ").replace("\r", " "))
        writer.writerow(clean_row)

print("CSV limpio generado correctamente:", csv_file)
